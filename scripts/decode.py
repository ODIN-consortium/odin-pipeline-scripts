import pandas as pd
import numpy as np
import os
import glob
import sys
import argparse
from openpyxl import load_workbook

# pathCurated = os.path.join('..', 'testdata')
# edna_curated_files = glob.glob(os.path.join(pathCurated, 'Synthetic*.xlsx'), recursive=True)
# assay_codes = pd.read_csv(os.path.join('..', 'config', 'assay-codes.txt'), sep='\t')
location_loc = 'B107'
location_tag_loc = 'B108'
lon_loc = 'B110'
lat_loc = 'B109'

def drp(df):  # Drop rows utility if Cq is 0
    if (df['Cq'] == 0).all():
        return df
    else:
        return df.drop(df.index[df['Cq'] == 0.00])


def extract_edna_cq_values(edna_ws): # -> dict:
    """
    Extract and return required data from Excel sheet
    :param edna_file: Filename of Excel sheet containing eDNA PCR results
    :return: date, (date is returned as a string, as the qPCR data only has
                    a string representation and not a proper date format.
                    Pandas is probably our best bet for interpreting the date correctly.)
             run_name, (str - name given to the run)
             wells, (list - numerical list of the wells analyzed)
             sample_id, (list - numerical list of the samples analyzed)
             channels, (list - list of strings indicating which fluorophore has been used)
             cq, (list - numerical list of the cq values obtained by the analysis)
    """
    # Read edna excel
    # wb = load_workbook(edna_file)
    ws = edna_ws
    runtype = ws['A12'].value
    run_name = "UNDEFINED"
    date = np.datetime64('NaT')

    wells = []
    sample_id = []
    channels = []
    cq = []
    target_id = []

    if runtype == 'Cq':  # Then this is not a melt file
        run_name = ws['B1'].value
        date = ws['B2'].value
        # lon = ws[lon_loc].value
        # lat = ws[lat_loc].value
        # location_tag = ws[location_tag_loc].value
        # location = ws[location_loc].value

        wells_r = ws['B7:AB7']
        channels_r = ws['B8:AB8']
        sample_r = ws['B10:AB10']
        target_id_r = ws['B11:AB11']
        cq_r = ws['B12:AB12']

        for cell in wells_r[0]:
            try:
                wells.append(int(cell.value))
            except (ValueError, TypeError):
                wells.append(0)

        for cell in sample_r[0]:
            sample_id.append(str(cell.value).strip())

        for cell in channels_r[0]:
            channels.append(str(cell.value).strip())

        for cell in target_id_r[0]:
            target_id.append(str(cell.value).strip())

        for cell in cq_r[0]:
            try:
                cq.append(float(cell.value))
            except (ValueError, TypeError):
                cq.append(0.0)

    cq_dict = {'Date': date,
               'Run Name': run_name,
               'Well Number': wells,
               'Biomeme replicate': sample_id,  # Renamed from 'Assay'
               'Fluorophore': channels,
               'Target ID': target_id,
               'Cq': cq
               # 'lon': lon,
               # 'lat': lat,
               # 'location': location,
               # 'location_tag': location_tag
               }

    # print(f'{run_name}, , {date}, {location}, {location_tag}, {lat}, {lon}')

    return cq_dict


def extract_edna(edna_file: str, metadata_df: pd.DataFrame, row_metadata: dict = None) -> pd.DataFrame:
    wb = load_workbook(edna_file)
    sheets_list = []

    # Load assay-params.csv from ../config relative to this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    assay_params_path = os.path.join(script_dir, '..', 'config', 'assay-params.csv')
    assay_params_df = pd.read_csv(assay_params_path)
    assay_params_df['_target_id_lower'] = assay_params_df['Target_ID'].astype(str).str.lower()

    for sheet in wb:
        cq_dict = extract_edna_cq_values(sheet)
        raw_edna = pd.DataFrame.from_dict(cq_dict)
        raw_edna['Date'] = pd.to_datetime(raw_edna['Date'], format='mixed', yearfirst=True, dayfirst=False)
        raw_edna['Fluorophore'] = raw_edna['Fluorophore'].astype(str)
        raw_edna['Well Number'] = raw_edna['Well Number'].astype(int)

        abundance_df = raw_edna.copy()
        abundance_df = abundance_df[
            abundance_df['Target ID'].notna() &
            (abundance_df['Target ID'].astype(str).str.strip() != '') &
            (abundance_df['Target ID'].astype(str).str.lower() != 'none')
        ]
        # Create a temporary lowercased column for merging
        abundance_df['_target_id_lower'] = abundance_df['Target ID'].astype(str).str.lower()

        if not abundance_df.empty:
            abundance_df['Assay Date'] = abundance_df['Date']
            abundance_df = abundance_df.groupby([
                'Biomeme replicate',
                'Well Number',
                'Fluorophore',
                'Run Name',
                'Assay Date',
                'Target ID',
                '_target_id_lower'
                ],
                as_index=False, group_keys=True).mean()

            abundance_df = abundance_df.merge(
                assay_params_df,
                left_on='_target_id_lower',
                right_on='_target_id_lower',
                how='left',
                suffixes=('', '_assay')
            )

            abundance_df['Target ID'] = abundance_df['Target_ID']

            valid = (
                (abundance_df['Cq'] > 0) &
                abundance_df['B'].notnull() &
                abundance_df['M'].notnull()
            )
            abundance_df['Abundance'] = 10 ** ((abundance_df.loc[valid, 'Cq'] - abundance_df.loc[valid, 'B']) / abundance_df.loc[valid, 'M'])

            # Compute Abundance diluted
            if 'dilution_factor' in abundance_df.columns:
                abundance_df['Abundance diluted'] = abundance_df.apply(
                    lambda row: row['Abundance'] * row['dilution_factor']
                    if pd.notna(row['dilution_factor']) else row['Abundance'],
                    axis=1
                )
            else:
                abundance_df['Abundance diluted'] = abundance_df['Abundance']

            abundance_df = abundance_df.drop(columns=['_target_id_lower'])

            # Map Biomeme_sample_ID to each row for correct grouping
            if metadata_df is not None and 'Biomeme_sample_ID' in metadata_df.columns:
                biomeme_id_map = {}
                for idx, meta_row in metadata_df.iterrows():
                    sample_ids = [s.strip() for s in str(meta_row['Biomeme_sample_ID']).split(',') if s.strip()]
                    for sid in sample_ids:
                        biomeme_id_map[sid] = meta_row['Biomeme_sample_ID']
                abundance_df['Biomeme_sample_ID_group'] = abundance_df['Biomeme replicate'].map(biomeme_id_map)
            else:
                abundance_df['Biomeme_sample_ID_group'] = abundance_df['Biomeme replicate']

            # Calculate mean Abundance for each Biomeme_sample_ID group, Target ID, Run Name, and Fluorophore where Cq > 0 and not 'Pos'
            mean_abundance = abundance_df[
                (abundance_df['Cq'] > 0) &
                (abundance_df['Biomeme replicate'] != 'Pos')
            ].groupby(
                ['Biomeme_sample_ID_group', 'Target ID', 'Run Name', 'Fluorophore']
            )['Abundance'].mean().reset_index().rename(columns={'Abundance': 'Mean Abundance'})
            abundance_df = abundance_df.merge(
                mean_abundance,
                left_on=['Biomeme_sample_ID_group', 'Target ID', 'Run Name', 'Fluorophore'],
                right_on=['Biomeme_sample_ID_group', 'Target ID', 'Run Name', 'Fluorophore'],
                how='left'
            )
            abundance_df = abundance_df.drop(columns=['Biomeme_sample_ID_group'])

            # Calculate logMeanAbundance using numpy.log1p, only if Mean Abundance > 0
            abundance_df['logMeanAbundance'] = abundance_df['Mean Abundance'].apply(
                lambda x: np.log1p(x) if pd.notna(x) and x > 0 else np.nan
            )

            # Add quality_flag column: 0 = not checked, 1 = good, 2 = questionable, 3 = bad, 4 = below NTC threshold
            abundance_df['quality_flag'] = 0  # default

            # --- NEW GROUPING LOGIC FOR QUALITY FLAG ---
            # Build mapping from Biomeme replicate to sample_ids group
            sample_id_to_group = {}
            if metadata_df is not None and 'Biomeme_sample_ID' in metadata_df.columns:
                for idx, meta_row in metadata_df.iterrows():
                    sample_ids = [s.strip() for s in str(meta_row['Biomeme_sample_ID']).split(',') if s.strip()]
                    group_id = ','.join(sorted(sample_ids))  # Use a string for group identifier
                    for sid in sample_ids:
                        sample_id_to_group[sid] = group_id
            # Assign group_id to each row (excluding NTC)
            abundance_df['sample_group'] = abundance_df['Biomeme replicate'].map(sample_id_to_group)
            not_ntc_mask = ~abundance_df['Biomeme replicate'].str.contains('NTC', case=False, na=False)
            # For each group+Target ID, count non-zero Cq values (excluding NTC)
            group_counts = abundance_df[not_ntc_mask & (abundance_df['Cq'] > 0)].groupby([
                'sample_group', 'Target ID'
            ])['Cq'].count().reset_index().rename(columns={'Cq': 'group_nonzero_cq_count'})
            abundance_df = abundance_df.merge(group_counts, on=['sample_group', 'Target ID'], how='left')
            # Assign quality flag
            abundance_df['quality_flag'] = 0  # default
            abundance_df.loc[abundance_df['group_nonzero_cq_count'] == 1, 'quality_flag'] = 2
            abundance_df.loc[abundance_df['group_nonzero_cq_count'] > 1, 'quality_flag'] = 1
            abundance_df.loc[abundance_df['Cq'] == 0, 'quality_flag'] = 3

            # --- NTC threshold logic (unchanged) ---
            ntc_mask = abundance_df['Biomeme replicate'].str.contains('NTC', case=False, na=False)
            ntc_rows = abundance_df[ntc_mask & (abundance_df['Cq'] > 0)]
            abundance_df['quality_flag'] = abundance_df['quality_flag'].astype(int)
            for _, ntc_row in ntc_rows.iterrows():
                fluor = ntc_row['Fluorophore']
                target = ntc_row['Target ID']
                ntc_cq = ntc_row['Cq']
                match_mask = (
                    (abundance_df['Fluorophore'] == fluor) &
                    (abundance_df['Target ID'] == target) &
                    (~abundance_df['Biomeme replicate'].str.contains('NTC', case=False, na=False)) &
                    (abundance_df['Cq'] > 0) &
                    (abundance_df['Cq'] < ntc_cq)
                )
                abundance_df.loc[match_mask, 'quality_flag'] = 4
            # Remove the drop statement here. Do NOT drop columns yet.

            # Add string column 'presence' with new logic:
            # Do not set any presence flag if the Biomeme replicate contains "NTC"
            abundance_df['presence'] = abundance_df.apply(
                lambda row: (
                    None if pd.notna(row['Biomeme replicate']) and 'NTC' in str(row['Biomeme replicate']).upper()
                    else (
                        'true' if row['quality_flag'] == 1 and pd.notna(row['Mean Abundance']) and row['Mean Abundance'] > 0
                        else 'true_x' if row['quality_flag'] == 2 and pd.notna(row['Mean Abundance']) and row['Mean Abundance'] > 0
                        else 'false'
                    )
                ),
                axis=1
            )
            # Remove only new grouping columns (drop once, after all logic)
            abundance_df = abundance_df.drop(columns=['group_nonzero_cq_count', 'sample_group'])

            # --- NEW LOGIC FOR METADATA ---
            if metadata_df is not None and 'biomeme_run_name' in metadata_df.columns and 'Biomeme_sample_ID' in metadata_df.columns:
                for idx, meta_row in metadata_df.iterrows():
                    run_name = meta_row['biomeme_run_name']
                    if pd.isna(meta_row['Biomeme_sample_ID']):
                        continue
                    sample_ids = [s.strip() for s in str(meta_row['Biomeme_sample_ID']).split(',') if s.strip()]
                    # Filter abundance_df for this run_name and sample_ids
                    filtered = abundance_df[
                        (abundance_df['Run Name'] == run_name) &
                        (abundance_df['Biomeme replicate'].isin(sample_ids))
                    ].copy()
                    # Attach metadata columns to filtered rows
                    for col in metadata_df.columns:
                        filtered[col] = meta_row[col]
                    sheets_list.append(filtered)
            else:
                sheets_list.append(abundance_df)
        else:
            sheets_list.append(abundance_df)
    abundance_list = pd.concat(sheets_list, ignore_index=True)
    # Remove abundance, abundance diluted, and mean abundance for all NTC rows
    ntc_mask = abundance_list['Biomeme replicate'].str.contains('NTC', case=False, na=False)
    abundance_list.loc[ntc_mask, ['Abundance', 'Abundance diluted', 'Mean Abundance']] = np.nan

    # # --- CORRECTED LOGIC: Merge site metadata from metadata_df ---
    # # Assume metadata_df contains the 'sites' information
    # if metadata_df is not None:
    #     # Look for site metadata columns in metadata_df
    #     site_cols = ['site', 'site_ID', 'location']
    #     # Find a suitable merge key
    #     merge_key = None
    #     for key in ['Biomeme replicate', 'Sample_ID', 'Biomeme_sample_ID', 'sample_id']:
    #         if key in abundance_list.columns and key in metadata_df.columns:
    #             merge_key = key
    #             break
    #     if not merge_key:
    #         # Fallback: try to merge on 'Biomeme replicate' <-> 'sample_id' if possible
    #         if 'Biomeme replicate' in abundance_list.columns and 'sample_id' in metadata_df.columns:
    #             merge_key = 'Biomeme replicate'
    #             metadata_df = metadata_df.rename(columns={'sample_id': 'Biomeme replicate'})
    #     if merge_key and all(col in metadata_df.columns for col in site_cols):
    #         abundance_list = abundance_list.merge(
    #             metadata_df[site_cols + [merge_key]],
    #             on=merge_key,
    #             how='left'
    #         )
    #     else:
    #         # If no merge key or site columns found, just add empty columns
    #         for col in site_cols:
    #             abundance_list[col] = np.nan
    # else:
    #     # If no metadata_df, add empty columns
    #     for col in ['site', 'site_ID', 'location']:
    #         abundance_list[col] = np.nan

    return abundance_list.copy()
